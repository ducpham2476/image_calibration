import openpyxl

def newWB():
  wb = openpyxl.Workbook()
  sheet = wb.active
  return wb, sheet

def process(sheet, data, lost_y):
  count = 1
  for i in range(len(data)):
    for j in range(len(data[i])):  
      sheet.cell(row=process.counter, column=count, value=data[i][j])
      count += 1
  for i in lost_y:
    if(i == 1):
      continue
    sheet.cell(row=process.counter, column=count, value=i)
    count += 1
  sheet.cell(row=process.counter, column=14, value=(len(lost_y) - 1))
  process.counter += 1
  return sheet

process.counter = 1


